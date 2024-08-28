# pip install pandas
# pip install openpyxl

import os, re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Protection
from openpyxl.cell import Cell
import copy

# Путь к папке с исходниками .csv
folder_path_csv = r"D:\UK\GitHub"
# Путь к папке для сохранения справочника
folder_path_sprav = r"D:\UK\GitHub"
# Путь к файлу с дополнениями
add_file_path = "D:/UK/GitHub/Additions.xlsx"
# Путь к целевому файлу справочника
target_file_path = "D:/UK/GitHub/Telephone_directory.xlsx"

# Получаем список файлов .csv в папке
files = [file for file in os.listdir(folder_path_csv) if file.endswith('.csv')]

# Определяем файл с последней датой изменения
latest_file = max(files, key=lambda f: os.path.getmtime(os.path.join(folder_path_csv, f)))

# Полный путь к последнему измененному файлу
latest_file_path = os.path.join(folder_path_csv, latest_file)

# Выводим имя выбранного файла
print(f"\nПоследний измененный файл: {latest_file}\n")

# Заголовки для объектов
headers = [
    "UserName", "AlternameName", "InternetAddress", "Title",
    "Company", "Department", "Manager", "OfficePhone",
    "CorporatePhone", "ExternalPhone"
]

# Словарь весов для каждого значения Title
title_weights = {
    "Управляющий директор ПАО \"Южный Кузбасс\"": 1,
    "Директор Департамента": 2,
    "Директор департамента - Главный бухгалтер": 3,
    "Директор разреза": 4,
    "Директор службы": 5,
    "Директор Управления": 6,
    "Директор учебного центра": 7,
    "Директор филиала": 8,
    "Директор шахты": 9,
    "Директор": 10,
    "Секретарь ГОФ \"Томусинская\"": 11,
    "Секретарь Красногорский разрез": 12,
    "Секретарь Ленина шахта": 13,
    "Секретарь Ольжерасская шахта": 14,
    "Секретарь Ольжерасский разрез": 15,
    "Секретарь ОФ \"Красногорская\"": 16,
    "Секретарь руководителя": 17,
    "Секретарь Сибиргинская шахта": 18,
    "Секретарь Сибиргинский разрез": 19,
    "Секретарь ТАУ ЮК": 20,
    "Секретарь УКК": 21,
    "Секретарь УОДУ ЮК": 22,
    "Секретарь УОиПУ ЮК": 23,
    "Секретарь УПДУ ЮК": 24,
    "Секретарь ЦОФ \"Кузбасская\"": 25,
    "Секретарь ЦОФ \"Сибирь\"": 26,
    "Секретарь ЮК": 27,
    "Технический директор": 28,
    "Технический директор - Директор департамента": 29,
    "Заместитель директора по производству": 30,
    "Заместитель директора по производству - Начальник цеха": 31,
    "Заместитель директора по производству - Начальник цеха (подземный)": 32,
    "Помощник директора": 33,
    "Руководитель направления": 34,
    "Руководитель направления по ремонтам и эксплуатации ДВС": 35,
    "Руководитель службы": 36,
    "Начальник автоколонны": 37,
    "Начальник гаража": 38,
    "Начальник группы": 39,
    "Начальник отдела": 40,
    "Начальник отделения": 41,
    "Начальник смены": 42,
    "Начальник смены Сибиргинский разрез": 43,
    "Начальник смены ЮК": 44,
    "Начальник участка": 45,
    "Начальник участка (подземный)": 46,
    "Начальник цеха": 47,
    "Начальник цеха - Главный механик": 48,
    "Заместитель начальника подземного участка": 49,
    "Заведующий складом": 50,
    "Заведующий складом взрывчатых материалов": 51,
    "Помощник начальника участка (подземный)": 52,
    "Председатель совета ветеранов": 53,
    "Главный экономист": 54,
    "Главный энергетик": 55,
    "Главный геолог": 56,
    "Главный геолог подземный": 57,
    "Главный геолог разреза": 58,
    "Главный инженер": 59,
    "Главный маркшейдер": 60,
    "Главный маркшейдер подземный": 61,
    "Главный маркшейдер разреза": 62,
    "Главный механик": 63,
    "Главный сварщик": 64,
    "Главный технолог": 65,
    "Главный специалист": 66,
    "Главный специалист по безопасности": 67,
    "Главный специалист по буровзрывным работам": 68,
    "Главный специалист по внешним и внутренним коммуникациям": 69,
    "Главный специалист по подбору и адаптации персонала": 70,
    "Энергетик": 71,
    "Заместитель главного инженера по производству подземный": 72,
    "Механик": 73,
    "Механик участка": 74,
    "Механик участка (подземный)": 75,
    "Сменный механик": 76,
    "Ведущий бухгалтер": 77,
    "Бухгалтер": 78,
    "Ведущий геолог": 79,
    "Ведущий гидрогеолог": 80,
    "Ведущий горный инженер": 81,
    "Ведущий инженеp-лабоpант": 82,
    "Ведущий инженер": 83,
    "Ведущий инженер по безопасности движения": 84,
    "Ведущий инженер по проектно-сметной работе": 85,
    "Ведущий инженер электросвязи": 86,
    "Ведущий инженер-системотехник": 87,
    "Горный инженер": 88,
    "Инженер": 89,
    "Инженер по комплектации оборудования": 90,
    "Инженер по организации управления производством": 91,
    "Инженер по проектно-сметной работе": 92,
    "Инженер электросвязи": 93,
    "Инженер-конструктор": 94,
    "Инженер-системотехник": 95,
    "Инженер-электроник": 96,
    "Ведущий маркшейдер": 97,
    "Ведущий программист": 98,
    "Ведущий системный администратор информационно-коммуникационных систем": 99,
    "Ведущий системный аналитик": 100,
    "Ведущий специалист": 101,
    "Ведущий специалист по безопасности": 102,
    "Ведущий специалист по закупкам": 103,
    "Ведущий специалист по землеустройству": 104,
    "Ведущий специалист по информационным системам": 105,
    "Ведущий специалист по качеству продукции": 106,
    "Ведущий специалист по лицензированию": 107,
    "Ведущий специалист по маркетингу": 108,
    "Ведущий специалист по оплате труда": 109,
    "Ведущий специалист по охране труда": 110,
    "Ведущий специалист по подбору и адаптации персонала": 111,
    "Ведущий специалист по пожарной безопасности, гражданской обороне и предупреждению чрезвычайных ситуаций": 112,
    "Ведущий специалист по промышленной безопасности": 113,
    "Ведущий специалист по прочему транспорту": 114,
    "Ведущий специалист по работе с персоналом": 115,
    "Ведущий специалист по радиосвязи и телекоммуникациям": 116,
    "Ведущий специалист по развитию и обучению персонала": 117,
    "Ведущий специалист по социальным программам": 118,
    "Ведущий специалист по технической поддержке": 119,
    "Ведущий специалист по экологической безопасности": 120,
    "Ведущий экономист": 121,
    "ведущий экономист по планированию": 122,
    "Ведущий юрисконсульт": 123,
    "Ведущий юрисконсульт по корпоративной работе": 124,
    "Юрисконсульт": 125,
    "Горный мастер": 126,
    "Горный мастер подземный": 127,
    "Бригадир на участках основного производства": 128,
    "Системный администратор информационно-коммуникационных систем": 129,
    "Старший диспетчер": 130,
    "Старший инженер-электроник": 131,
    "Старший кладовщик": 132,
    "Старший лаборант химического анализа": 133,
    "Старший мастер": 134,
    "Старший механик": 135,
    "старший механик по автоматике подземный": 136,
    "Старший механик по буровым станкам ": 137,
    "Старший приемосдатчик груза и багажа": 138,
    "Старший энергетик": 139,
    "Мастер": 140,
    "Мастер буровой": 141,
    "Мастер дорожный": 142,
    "Мастер контрольный": 143,
    "мастер ОТК р-з Красногорский": 144,
    "мастер ОТК ЦОФ \"Кузбасская\"": 145,
    "Мастер по ремонту оборудования": 146,
    "Мастера ОТК разрез Сибиргинский": 147,
    "Аппаратчик углеобогащения": 148,
    "Архивариус": 149,
    "Взрывник подземный": 150,
    "Водитель погрузчика": 151,
    "Геолог": 152,
    "Горный диспетчер": 153,
    "Дефектоскопист по ультразвуковому контролю": 154,
    "Диспетчер": 155,
    "Кладовщик": 156,
    "Комендант": 157,
    "Контролер материалов, металлов, полуфабрикатов и изделий": 158,
    "Контролер по учету и контролю расхода ГСМ": 159,
    "Контролер продукции обогащения": 160,
    "Контролер технического состояния автомототранспортных средств": 161,
    "Контролер-учетчик": 162,
    "Лаборант химического анализа": 163,
    "Ламповщик": 164,
    "Маркшейдер": 165,
    "Машинист компрессорных установок": 166,
    "Машинист насосных установок": 167,
    "Оператор автоматической газовой защиты": 168,
    "Оператор ЭВ и ВМ": 169,
    "Оператор электронно-вычислительных и вычислительных машин": 170,
    "Приемосдатчик груза и багажа": 171,
    "Программист": 172,
    "Рабочий по комплексному обслуживанию зданий": 173,
    "Системный аналитик": 174,
    "Слесарь по обслуживанию и ремонту оборудования": 175,
    "Слесарь по ремонту автомобилей": 176,
    "Слесарь-ремонтник": 177,
    "Специалист": 178,
    "Специалист по безопасности": 179,
    "Специалист по внешним и внутренним коммуникациям": 180,
    "Специалист по горюче-смазочным материалам": 181,
    "Специалист по закупкам": 182,
    "Специалист по землеустройству": 183,
    "Специалист по информационным системам": 184,
    "Специалист по кадровому администрированию": 185,
    "Специалист по качеству продукции": 186,
    "Специалист по лицензированию": 187,
    "Специалист по мобилизационной подготовке экономики": 188,
    "Специалист по охране труда": 189,
    "Специалист по подбору и адаптации персонала": 190,
    "Специалист по пожарной безопасности, гражданской обороне и предупреждению чрезвычайных ситуаций": 191,
    "Специалист по промышленной безопасности": 192,
    "Специалист по работе с инвестиционными проектами": 193,
    "Специалист по работе с персоналом": 194,
    "Специалист по развитию и обучению персонала": 195,
    "Специалист по социальным программам": 196,
    "Специалист по технической поддержке": 197,
    "Специалист по экологической безопасности": 198,
    "Специалист пропускного режима": 199,
    "Техник-картограф": 200,
    "Товаровед": 201,
    "Участковый геолог": 202,
    "Участковый геолог подземный": 203,
    "Участковый маркшейдер": 204,
    "Участковый маркшейдер подземный": 205,
    "Учетчик": 206,
    "Экономист": 207,
    "Электромеханик": 208,
    "Электромеханик по испытанию и ремонту электрооборудования": 209,
    "Электромонтер линейных сооружений телефонной связи и радиофикации": 210,
    "Электрослесарь по обслуживанию и ремонту оборудования": 211,
    "Электрослесарь подземный": 212,
    "Горнорабочий": 213,
    "Горнорабочий на маркшейдерских работах": 214,
    "Горнорабочий подземный": 215,
}

try:
    # Считываем количество строк в CSV-файле
    with open(latest_file_path, 'r', encoding='utf-8') as file:
        line_count = sum(1 for line in file)
    print(f"Количество строк в файле: {line_count}")

    # Читаем все строки файла в массив, пропуская первую строку
    with open(latest_file_path, 'r', encoding='utf-8') as file:
        values = file.readlines()[1:]
    
    # Массив объектов для хранения данных
    objects = []
    no_department_objects = []

    # Перебираем массив построчно
    for value in values:

        # Разделяем строку по разделителю
        parts = value.split(";")

        # Чистим пробелы в начале и конце каждой части
        parts = [part.strip() for part in parts]

        # Чистим UserName
        parts[0] = parts[0].split("/", 1)[0]
        if parts[0][0] == "\"":  # Удаляем первую кавычку
            parts[0] = parts[0][1:]

        # Чистим AlternameName
        parts[1] = parts[1].split("/", 1)[0]

        # Проверяем и заменяем строку перед сплитом
        if "ЭШ-20/90" in parts[5]:
            parts[5] = parts[5].replace("ЭШ-20/90", "ЭШ-20|90")

        # Переворачиваем Department для правильной сортировки
        dep_tmp = parts[5].split("/")
        dep_tmp.reverse()
        parts[5] = "/".join(dep_tmp)

        # Проверяем, содержит ли parts[5] "/0400" и обрезаем строку, оставляя "0400" и все, что после
        if "/0400" in parts[5]:
            index = parts[5].index("/0400")
            parts[5] = parts[5][index + 1:]  # Оставляем "0400" и все, что после него
            parts[5] = parts[5].replace("0400.  ", "0400. ")  # Убираем лишний пробел
            
        # Чистим Manager
        parts[6] = parts[6].split("/", 1)[0]

        # Правим OfficePhone (здесь - сотовый или МГ код)
        if parts[7]:
            if "38475" in parts[7]:
                parts[7] = parts[7].replace("38475", " (38475) ")
            else:
                if "38474" in parts[7]:
                    parts[7] = parts[7].replace("38474", " (38474) ")
                elif len(parts[7]) <= 12:
                    parts[7] = parts[7][:-10] + " " + parts[7][-10:-7] + "-" + parts[7][-7:-4] + "-" + parts[7][-4:-2] + "-" + parts[7][-2:]

        # Правим CorporatePhone
        if parts[8] != '':
            parts[8] = parts[8][:1] + "-" + parts[8][1:3] + "-" + parts[8][3:]

        # Правим ExternalPhone (здесь - 4х значный офисный номер)
        parts[9] = parts[9].strip()  # Убираем последние пробелы и символы перевода строки

        # Если последний символ кавычка, удаляем его и очищаем пробелы
        if parts[9] and parts[9][-1] == "\"":
            parts[9] = parts[9][:-1].strip()
        if parts[9] and len(parts[9]) > 5:  # Если поле не пустое и длина строки больше 5 символов
            parts[9] = parts[9][:2] + "-" + parts[9][2:-2] + "-" + parts[9][-2:]
        elif parts[9]:
            parts[9] = parts[9][:-2] + "-" + parts[9][-2:]


        # Если ExternalPhone пустой, берем данные из CorporatePhone, если и тот пустой - из OfficePhone
        if not parts[9]:
            parts[9] = parts[8].strip() if parts[8] else parts[7].strip()

        # Проверяем наличие четырех цифр подряд в начале строки в поле Department
        department_field = parts[5]

        # Ищем четыре цифры подряд в начале строки или в середине строки
        if not re.search(r'\b\d{4}\b', department_field) and not "29." in department_field:
            no_department_objects.append(dict(zip(headers, parts)))
        else:
            # Создаем объект с соответствующими ключами и значениями
            if len(parts) == len(headers):  # Убедимся, что количество значений совпадает с количеством заголовков
                obj = dict(zip(headers, parts))
                objects.append(obj)

    # Количество объектов в массиве
    object_count = len(objects)
    print(f"Количество объектов в массиве: {object_count}")

    # Отфильтруем строки с пустым полем Department
    no_department_objects.extend([obj for obj in objects if not obj.get("Department", "").strip()])
    with_department_objects = [obj for obj in objects if obj.get("Department", "").strip()]

    # Функция для получения веса Title
    def get_title_weight(title):
        return title_weights.get(title, 0)  # Возвращает 0, если title не найден в словаре

    # Функция для получения приоритета Department
    def department_priority(department):
        if "0100.01. Администрация ПАО \"Южный Кузбасс\"" in department:
            return 0  # Высокий приоритет для "0100.01. Администрация ПАО \"Южный Кузбасс\""
        elif department == "0100. Администрация":
            return 1  # Средний приоритет для "0100. Администрация"
        else:
            return 2  # Низкий приоритет для всех остальных департаментов

    # Сортировка объектов: сначала по Department с учетом приоритета, затем по Title с учетом веса
    with_department_objects.sort(
        key=lambda x: (department_priority(x.get("Department", "")), x.get("Department", ""), get_title_weight(x.get("Title", "")))
    )

    # Функция для вставки копий строк перед строками, начинающимися с определенного департамента
    def insert_before_department(with_department_objects, target_department, startswith_department):
        # Ищем строки с нужным Department
        target_objects = [obj for obj in with_department_objects if obj.get("Department", "") == target_department]

        # Ищем индекс первой строки, начинающейся с startswith_department
        insert_index = next((i for i, obj in enumerate(with_department_objects) if obj.get("Department", "").startswith(startswith_department)), None)

        # Вставляем найденные строки перед первой строкой, начинающейся с startswith_department
        if insert_index is not None:
            with_department_objects[insert_index:insert_index] = target_objects
        else:
            # Если не нашли, добавляем строки в конец
            with_department_objects.extend(target_objects)

    # Вставка для "0200. Управление по экономике и финансам"
    insert_before_department(
        with_department_objects,
        "0100.01. Администрация ПАО \"Южный Кузбасс\"/0200. Управление по экономике и финансам",
        "0200. Управление по экономике и финансам"
    )

    # Вставка для "0400. Управление по операционной деятельности"
    insert_before_department(
        with_department_objects,
        "0100.01. Администрация ПАО \"Южный Кузбасс\"/0400. Управление по операционной деятельности",
        "0400. Управление по операционной деятельности"
    )

    # Вставка для "0600. Управление по работе с персоналом"
    insert_before_department(
        with_department_objects,
        "0100.01. Администрация ПАО \"Южный Кузбасс\"/0600. Управление по работе с персоналом",
        "0600. Управление по работе с персоналом"
    )

    # Вставка для "0700. Управление корпоративной безопасности"
    insert_before_department(
        with_department_objects,
        "0100.01. Администрация ПАО \"Южный Кузбасс\"/0700. Управление корпоративной безопасности",
        "0700. Управление корпоративной безопасности"
    )

    # Создаем DataFrame из объектов с нужными колонками и добавляем столбец Department
    filtered_df = pd.DataFrame(with_department_objects, columns=["Title", "AlternameName", "ExternalPhone", "Department"])
    no_department_df = pd.DataFrame(no_department_objects, columns=["Title", "AlternameName", "ExternalPhone", "Department"])

    output_file_path = os.path.join(folder_path_sprav, "Telephone_directory.xlsx")
    with pd.ExcelWriter(output_file_path, engine="openpyxl") as writer:
        filtered_df.to_excel(writer, index=False, header=False, sheet_name="ЮК")  # Лист "ЮК"
        no_department_df.to_excel(writer, index=False, header=False, sheet_name="NoDepartment")  # Лист "NoDepartment"

    # Открываем созданный файл для форматирования
    wb = load_workbook(output_file_path)
    ws_with_department = wb["ЮК"]
    ws_no_department = wb["NoDepartment"]

    # Вставляем две строки в начало листа "ЮК"
    ws_with_department.insert_rows(1, 2)

    # Объединяем ячейки A1:C1 и A2:C2
    ws_with_department.merge_cells('A1:C1')
    ws_with_department.merge_cells('A2:C2')

    # Настраиваем форматирование для первой строки
    header_font = Font(name="Times New Roman", size=36, bold=True, color="0000FF")  # Синий цвет
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Заполняем первую строку текстом и применяем форматирование
    ws_with_department['A1'].value = "Телефонный справочник"
    ws_with_department['A1'].font = header_font
    ws_with_department['A1'].alignment = alignment_center

    # Заполняем вторую строку текстом и применяем форматирование
    ws_with_department['A2'].value = 'ПАО "Южный Кузбасс"'
    ws_with_department['A2'].font = header_font
    ws_with_department['A2'].alignment = alignment_center

    # Закрепляем первые две строки вверху страницы
    ws_with_department.freeze_panes = 'A4'

    # Форматирование столбцов (начиная с 3-й строки)
    column_formats = {
        "A": {"width": 50, "font": Font(name="Times New Roman", size=10, bold=True, italic=True), "wrap_text": True},  # Title with wrap text
        "B": {"width": 62, "font": Font(name="Times New Roman", size=12, bold=True)},  # AlternameName
        "C": {"width": 30, "font": Font(name="Times New Roman", size=14, bold=True), "alignment": Alignment(horizontal="center", vertical="center")},  # ExternalPhone
        "D": {"width": 30, "font": Font(size=10, color="FFFFFF"), "alignment": Alignment(vertical="center")}  # Department (No formatting needed)
    }

    for col, fmt in column_formats.items():
        # Устанавливаем ширину столбца для листа с Department
        ws_with_department.column_dimensions[col].width = fmt["width"]

        # Применяем шрифт и выравнивание к ячейкам столбца, начиная с третьей строки
        if "font" in fmt:
            for cell in ws_with_department[col][2:]:  # Начинаем с 3-й строки, так как 1 и 2 заняты заголовками
                cell.font = fmt["font"]
                if col in ["A", "B", "D"]:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=fmt.get("wrap_text", False))
                elif col == "C":
                    cell.alignment = fmt.get("alignment", Alignment(horizontal="center", vertical="center"))
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Устанавливаем ширину столбца для листа без Department
        ws_no_department.column_dimensions[col].width = fmt["width"]

        # Применяем шрифт и выравнивание к ячейкам столбца, если указано
        if "font" in fmt:
            for cell in ws_no_department[col]:
                cell.font = fmt["font"]
                if col in ["A", "B"]:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=fmt.get("wrap_text", False))
                elif col == "C":
                    cell.alignment = fmt.get("alignment", Alignment(horizontal="center", vertical="center"))
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    def process_departments_and_insert_rows(sheet, department_column_index):
        """
        Функция обрабатывает департаменты и вставляет строку заголовка перед строкой,
        где изменяется департамент, исключая первый случай указанных департаментов.
        
        :param sheet: Лист Excel, который необходимо обработать.
        :param department_column_index: Индекс колонки, в которой находится департамент.
        """
        previous_department = None

        # Определяем департаменты, для которых нужно исключить первую вставку
        exclude_departments = {
            "0100.01. Администрация ПАО \"Южный Кузбасс\"/0200. Управление по экономике и финансам",
            "0100.01. Администрация ПАО \"Южный Кузбасс\"/0600. Управление по работе с персоналом",
            "0100.01. Администрация ПАО \"Южный Кузбасс\"/0700. Управление корпоративной безопасности"
        }

        # Множество для отслеживания уже обработанных департаментов
        seen_departments = set()

        # Список для хранения строк, где нужно вставить заголовки
        rows_to_insert = []

        # Проходим по строкам
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            current_department = row[department_column_index - 1]  # Получаем текущий департамент из строки

            # Проверяем, нужно ли исключить вставку заголовка для первого появления департамента
            if current_department != previous_department and current_department is not None:
                if current_department in exclude_departments:
                    if current_department not in seen_departments:
                        seen_departments.add(current_department)
                    else:
                        # Если департамент уже встречался, добавляем индекс для вставки заголовка
                        rows_to_insert.append(row_index)
                else:
                    # Если департамент не в списке исключений, добавляем индекс для вставки заголовка
                    rows_to_insert.append(row_index)

            # Обновляем предыдущий департамент
            previous_department = current_department

        # Прямой порядок вставки с пересчетом индексов
        for i, insert_index in enumerate(rows_to_insert):
            # Пересчитываем индекс в зависимости от количества уже вставленных строк
            corrected_index = insert_index + i

            # Вставляем строку перед текущей строкой
            sheet.insert_rows(corrected_index)

            # Определяем текст для объединенной ячейки
            # Нужно получить департамент из строки ниже, которая сместится
            current_department = sheet.cell(row=corrected_index + 1, column=department_column_index).value
            if current_department is None:
                current_department = ""
            
            if '/' in current_department:
                text_to_insert = current_department.split('/')[-1].strip()
            else:
                text_to_insert = current_department

            # Объединяем ячейки A, B, и C в новой строке
            sheet.merge_cells(start_row=corrected_index, start_column=1, end_row=corrected_index, end_column=3)

            # Вставляем значение в объединенную ячейку
            merged_cell = sheet.cell(row=corrected_index, column=1)
            merged_cell.value = text_to_insert

            # Форматирование для вставленной строки
            font_style = Font(name="Times New Roman", size=14, bold=True, italic=True, color="000080")  # Синий курсив жирный
            alignment_style = Alignment(horizontal="center", vertical="center")  # Выравнивание по центру
            fill_style = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Светло-зеленый цвет

            # Применяем форматирование к объединенной ячейке
            merged_cell.font = font_style
            merged_cell.alignment = alignment_style

            # Применяем цвет фона только к ячейкам A, B и C
            for col in range(1, 4):  # Колонки A (1), B (2), C (3)
                cell = sheet.cell(row=corrected_index, column=col)
                cell.fill = fill_style

        print(f"Добавлено {len(rows_to_insert)} заголовков.")

    process_departments_and_insert_rows(ws_with_department, department_column_index=4)

    # Сохраняем форматированный файл и с новым именем листа
    wb.save(output_file_path)

    print(f"\nДанные сохранены и отформатированы в файл: {output_file_path}")

except Exception as e:
    print(f"Ошибка при чтении файла: {e}")


# ДОБАВЛЕНИЕ ДОПОЛНЕНИЙ

# Загрузка целевого файла
wb_target = load_workbook(target_file_path)
ws_target = wb_target.active

# Загрузка файла с дополнениями
wb_add = load_workbook(add_file_path)
ws_add = wb_add.active

# Подсчет количества строк в файле дополнений
row_count_add = ws_add.max_row

# Сохраняем текущие данные и форматирование целевого файла начиная с 3-й строки
rows_data = []
for row in ws_target.iter_rows(min_row=3, max_row=ws_target.max_row):
    row_data = []
    for cell in row:
        if isinstance(cell, Cell):  # Проверяем, что ячейка не объединённая
            cell_data = {
                "value": cell.value,
                "font": copy.copy(cell.font),
                "alignment": copy.copy(cell.alignment),
                "fill": copy.copy(cell.fill),
                "border": copy.copy(cell.border),
                "protection": copy.copy(cell.protection),
            }
            row_data.append(cell_data)
        else:
            row_data.append(None)  # Для объединённых ячеек добавляем None
    rows_data.append(row_data)

# Сохраняем информацию о текущих объединённых диапазонах
merged_ranges = list(ws_target.merged_cells.ranges)

# Разъединяем ячейки, если они пересекаются с диапазоном вставки
merged_cells_to_remove = []
for merged_range in merged_ranges:
    min_row, max_row = merged_range.min_row, merged_range.max_row
    if min_row >= 3:
        merged_cells_to_remove.append(merged_range)

for merged_range in merged_cells_to_remove:
    ws_target.unmerge_cells(str(merged_range))

# Вставляем пустые строки в целевой файл, начиная с 3-й строки
ws_target.insert_rows(idx=3, amount=row_count_add)

# Вставляем сохраненные данные и форматирование на новые позиции
for row_idx, row_data in enumerate(rows_data, start=3 + row_count_add):
    for col_idx, cell_data in enumerate(row_data, start=1):
        target_cell = ws_target.cell(row=row_idx, column=col_idx)
        if cell_data is not None:
            target_cell.value = cell_data["value"]
            target_cell.font = cell_data["font"]
            target_cell.alignment = cell_data["alignment"]
            target_cell.fill = cell_data["fill"]
            target_cell.border = cell_data["border"]
            target_cell.protection = cell_data["protection"]

# Копируем данные и форматирование из файла дополнений в целевой файл
for row_idx, row in enumerate(ws_add.iter_rows(min_row=1, max_row=row_count_add, values_only=False), start=3):
    for col_idx, cell in enumerate(row, start=1):
        target_cell = ws_target.cell(row=row_idx, column=col_idx)
        if isinstance(cell, Cell):  # Проверяем, что ячейка не объединённая
            target_cell.value = cell.value
            target_cell.font = copy.copy(cell.font)
            target_cell.alignment = copy.copy(cell.alignment)
            target_cell.fill = copy.copy(cell.fill)
            target_cell.border = copy.copy(cell.border)
            target_cell.protection = copy.copy(cell.protection)

# Восстанавливаем объединённые ячейки
for merged_range in merged_ranges:
    # Учитываем возможное смещение строк
    if merged_range.min_row >= 3:
        new_start_row = merged_range.min_row + row_count_add
        new_end_row = merged_range.max_row + row_count_add
        ws_target.merge_cells(start_row=new_start_row, start_column=merged_range.min_col,
                              end_row=new_end_row, end_column=merged_range.max_col)

# Сохранение изменений в целевом файле
wb_target.save(target_file_path)

# Вывод количества вставленных строк на экран
print(f"Количество вставленных строк: {row_count_add}")
